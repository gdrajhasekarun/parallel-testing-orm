import ast
import datetime
import io
import itertools
import uuid
from typing import List

from openpyxl import load_workbook
from sqlalchemy import MetaData, distinct, and_, case
from sqlalchemy import insert, inspect, Table, select, func, or_
from starlette.responses import StreamingResponse

from app.database import engine, SessionLocal, update_tables
from app.models.model import job_store
# from app.tables.dynamic_tables import create_data_table
from app.tables.tables import ParallelData, CompareSession, Differences
import pandas as pd


class ExcelHandler:
    BATCH_SIZE = 10000

    def __init__(self):
        self.metadata = MetaData()
        self.inspector = inspect(engine.engine)
        self.session = SessionLocal()

    def _process_sheet(self, work_sheet, sheet_name: str, source: str, date: datetime.date):
        header = None
        batch = []
        table = None

        for i, row in enumerate(work_sheet.iter_rows(values_only=True), start=1):
            row = ['Null' if cell is None else str(cell) for cell in row]
            if i == 1:
                header = row
                # if not header or all(col == 'Null' for col in header):
                #     print(f"Warning: Sheet '{sheet_name}' is empty. Skipping...")
                #     return
                #
                # # Create or get the dynamic SQLAlchemy table
                # table = create_data_table(f"data_{update_tables(sheet_name)}", header, self.metadata)
                # table.create(bind=engine.engine, checkfirst=True)
                continue

            row += ['Null'] * (len(header) - len(row))
            row = row[:len(header)]
            row_dict = dict(zip(header, row))
            row_dict['d_data_source'] = source
            row_dict['d_sheet'] = sheet_name
            row_dict['d_date'] = date
            batch.append(row_dict)

            if len(batch) >= self.BATCH_SIZE:
                # self.session.execute(insert(table), batch)
                self.session.execute(insert(ParallelData), batch)
                self.session.commit()
                batch = []

        if batch:
            # self.session.execute(insert(table), batch)
            self.session.execute(insert(ParallelData), batch)
            self.session.commit()

        print(f"Sheet '{sheet_name}' processed successfully.")

    def excel_to_db(self, excel_file_content, source:str, date: datetime.date):
        try:
            wb = load_workbook(excel_file_content, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                self._process_sheet(wb[sheet_name], sheet_name, source, date)
            wb.close()
        except Exception as e:
            raise Exception(f"Failed to process the file: {e}")

    def get_list_tables(self, data_source: str, date: datetime.date) -> List[str]:
        result = self.session.query(
            distinct(ParallelData.d_sheet)
        ).filter(
            and_(
                func.lower(ParallelData.d_data_source) == data_source.lower(),
                func.date(ParallelData.d_date) == date
            )
        ).all()
        if result:
            return [row[0] for row in result]
        return []

    def get_col_names_from_db(self, sheet_name: str) -> List[str]:
        # columns_info = self.inspector.get_columns(update_tables(sheet_name))
        # column_names = [col['name'] for col in columns_info]
        # return column_names
        return  [column.name for column in ParallelData.__table__.columns if not column.name.startswith("d_")]

    # def create_batch_job_setup(self, db_name: str, table_name: str, primary_col: List[str]):
    def create_batch_job_setup(self, primary_col: List[str]):
        # primary_col.insert(0, "batch_id")
        # batch_table = create_data_table(f"data_{update_tables(table_name)}", primary_col, self.metadata)
        # batch_table.create(bind=engine, checkfirst=True)
        # return self.__generate_unique_combinations(batch_table, primary_col)
        return self.__generate_unique_combinations(primary_col)

    # def __generate_unique_combinations(self, batch_table: Table, primary_cols: List[str]):
    def __generate_unique_combinations(self, primary_cols: List[str]):
        # source_table = self.metadata.tables.get(batch_table.name)
        # if not source_table:
        #     raise Exception(f"Table {batch_table.name} not found in source DB")
        # col_exprs = [source_table.c[col] for col in primary_cols]
        # stmt = select(*col_exprs).distinct()
        # results = self.session.execute(stmt).fetchall()
        # combinations = list(itertools.combinations(results, len(primary_cols)))

        distinct_values = {}
        for col in primary_cols:
            stmt = select(ParallelData.__table__.c[col]).distinct()
            values = [row[0] for row in self.session.execute(stmt).fetchall()]
            distinct_values[col] = values

        combinations = [dict(zip(primary_cols, values)) for values in itertools.product(*distinct_values.values())]
        # Generate batch ID and insert combinations into batch table
        batch_id = str(uuid.uuid4())
        for combo in combinations:
            # flattened = [item if isinstance(item, str) else item[0] for item in combo]
            # row = {"batch_id": batch_id, **dict(zip(primary_cols, flattened)), "result": ""}
            unique_str = str(combo)
            row = {"batch_id": batch_id, "unique_value": unique_str, "result": "", "created_date": datetime.date.today()}
            self.session.execute(insert(CompareSession).values(row))
            self.session.commit()
        return {"jobId": batch_id, "total_combinations": len(combinations)}

    def get_status_job_id(self, job_id: str):
        stmt = select(
            func.count(case((CompareSession.result == 'Pass', 1))).label('pass_count'),
            func.count(case((CompareSession.result == 'Fail', 1))).label('fail_count'),
            func.count(case((
                or_(
                    CompareSession.result == None,
                    CompareSession.result == ''
                ), 1
            ))).label('blank_count')
        ).where(CompareSession.batch_id == job_id)

        result = self.session.execute(stmt).one()
        pass_count, fail_count, blank_count = result.pass_count, result.fail_count, result.blank_count
        if blank_count == 0:
            job_store[job_id] = "completed"
        return pass_count, fail_count, blank_count

    def _get_all_data_from_table_orm(self, model, filter_criteria):
        # filter_criteria can be a list/tuple of conditions or a single and_() condition
        # If filter_criteria is a single and_() condition, unpack it with *filter_criteria.clauses
        if hasattr(filter_criteria, 'clauses'):
            filters = list(filter_criteria.clauses)
        elif isinstance(filter_criteria, (list, tuple)):
            filters = list(filter_criteria)
        else:
            filters = [filter_criteria]

        stmt = select(model).where(*filters)
        results = self.session.execute(stmt).scalars().all()
        dict_list = [obj.__dict__.copy() for obj in results]
        for d in dict_list:
            d.pop('_sa_instance_state', None)
        return pd.DataFrame(dict_list)

    def export_to_excel(self, batch_id: str):
        filters = [CompareSession.batch_id == batch_id]
        result = self._get_all_data_from_table_orm(CompareSession, filters)
        differences = self._get_all_data_from_table_orm(Differences, [Differences.batch_id == batch_id])
        combined_df = result
        if len(differences) > 0:
            combined_df = pd.merge(result, differences, how='left', left_on='id', right_on='run_id')
            # Drop the `run_id` column if you don't want it in the final result
            combined_df.drop('run_id', axis=1, inplace=True)

        combined_df['unique_dict'] = combined_df['unique_value'].apply(ast.literal_eval)
        expanded_df = combined_df['unique_dict'].apply(pd.Series)
        combined_df = pd.concat([combined_df.drop(columns=['unique_value', 'unique_dict', 'id', 'batch_id']), expanded_df], axis=1)
        cols = [col for col in combined_df.columns if col != 'result'] + ['result']
        combined_df = combined_df[cols]
        # Export to Excel
        # combined_df.to_excel('Result.xlsx', index=False)
        # Save DataFrame to an in-memory buffer
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            combined_df.to_excel(writer, index=False, sheet_name="Combined Data")

        output.seek(0)  # Reset buffer position to the beginning

        # Return file as a response
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=Result.xlsx"})